# report_export.py — Funções de exportação PDF e Excel para relatórios

import os
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Carrega configurações
from dotenv import load_dotenv
env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
load_dotenv(env_path, override=True)

SCHOOL_NAME = os.getenv('SCHOOL_NAME', 'Biblioteca')
SISTEMA_NAME = 'LUMEN'

# ═══════════════════════════════════════════════════════════════════════════════
# PALETA DE CORES INSTITUCIONAL
# ═══════════════════════════════════════════════════════════════════════════════
COR_AZUL_MARINHO = (11, 29, 52)      # #0B1D34
COR_DOURADO = (201, 162, 76)         # #C9A24C
COR_BRANCO = (255, 255, 255)
COR_BEGE = (250, 247, 242)           # #FAF7F2
COR_CINZA_LINHA = (220, 215, 210)
COR_CINZA_TEXTO = (100, 100, 100)
COR_VERMELHO_SUAVE = (245, 230, 230)

# Ícones para relatórios (ASCII compatible)
ICONES_RELATORIOS = {
    'inventario': '[I]',
    'disponiveis': '[OK]',
    'emprestimos': '[>>]',
    'atrasos': '[!]',
    'historico_usu': '[U]',
    'mais_emp': '[#1]',
    'top_leit': '[*]',
    'aquisicoes': '[+]',
}

# ═══════════════════════════════════════════════════════════════════════════════
# CLASSE PDF PROFISSIONAL
# ═══════════════════════════════════════════════════════════════════════════════

class RelatorioPDF(FPDF):
    """PDF profissional estilo ERP com cabeçalho institucional."""

    def __init__(self, titulo_relatorio='', descricao='', icone=''):
        super().__init__(orientation='L', format='A4')
        self._titulo_relatorio = titulo_relatorio
        self._descricao = descricao
        self._icone = icone
        self._logo_escola = None
        self._logo_lumen = None
        self._carregar_logos()
        self.set_auto_page_break(auto=True, margin=25)

    def _carregar_logos(self):
        """Carrega as logos do sistema."""
        assets_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets')

        logo_escola = os.path.join(assets_dir, 'logo_escola.png')
        if os.path.exists(logo_escola):
            self._logo_escola = logo_escola

        logo_lumen = os.path.join(assets_dir, 'logo_lumen.png')
        if os.path.exists(logo_lumen):
            self._logo_lumen = logo_lumen

    def header(self):
        """Cabeçalho profissional com logos e informações institucionais."""

        # ═══ 1º: DESENHAR BANNER COMO FUNDO (se houver) ═══
        # Calcula posição Y onde o banner ficará
        if self._titulo_relatorio:
            # Pré-calcula a posição Y do banner baseado no espaço do texto
            y_banner = 8 + 9 + 5 + 5 + 4  # set_y(8) + cell(9) + cell(5) + cell(5) + ln(4) = 31
            # Desenha fundo azul-marinho do banner
            self.set_fill_color(*COR_AZUL_MARINHO)
            self.rect(12, y_banner, 273, 22, 'F')
            # Borda dourada inferior
            self.set_draw_color(*COR_DOURADO)
            self.set_line_width(0.5)
            self.line(12, y_banner + 22, 285, y_banner + 22)

        # ═══ 2º: TEXTOS CENTRALIZADOS ═══
        self.set_y(8)

        # Nome da escola (destaque)
        self.set_font('Helvetica', 'B', 16)
        self.set_text_color(*COR_AZUL_MARINHO)
        self.cell(0, 9, SCHOOL_NAME, new_x="LMARGIN", new_y="NEXT", align='C')

        # Subtítulo do sistema
        self.set_font('Helvetica', '', 10)
        self.set_text_color(*COR_CINZA_TEXTO)
        self.cell(0, 5, f'Sistema {SISTEMA_NAME} - Relatorios da Biblioteca',
                  new_x="LMARGIN", new_y="NEXT", align='C')

        # Data e hora
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 5, f'Gerado em: {datetime.now().strftime("%d/%m/%Y as %H:%M")}',
                  new_x="LMARGIN", new_y="NEXT", align='C')

        self.ln(4)

        # ═══ 3º: DESENHAR TEXTOS DO BANNER (se houver) ═══
        if self._titulo_relatorio:
            self._desenhar_banner_conteudo()

        # ═══ 4º: LOGOS POR CIMA DE TUDO ═══
        if self._logo_lumen:
            try:
                self.image(self._logo_lumen, x=14, y=10, w=18)
            except:
                pass

        if self._logo_escola:
            try:
                self.image(self._logo_escola, x=260, y=6, w=26)
            except:
                pass

    def _desenhar_banner_conteudo(self):
        """Desenha apenas o conteúdo textual do banner (fundo já desenhado no header)."""
        y_inicio = self.get_y()

        # Ícone (se houver)
        x_icone = 18
        if self._icone:
            self.set_font('Helvetica', '', 14)
            self.set_text_color(*COR_DOURADO)
            self.set_xy(x_icone, y_inicio + 4)
            self.cell(10, 14, self._icone, align='L')

        # Titulo do relatorio
        self.set_font('Helvetica', 'B', 13)
        self.set_text_color(*COR_BRANCO)
        x_titulo = 30 if self._icone else 18
        self.set_xy(x_titulo, y_inicio + 2)
        self.cell(240, 8, f'RELATORIO: {self._titulo_relatorio.upper()}', align='L')

        # Descrição
        if self._descricao:
            self.set_font('Helvetica', '', 8)
            self.set_text_color(180, 190, 210)
            self.set_xy(x_titulo, y_inicio + 11)
            self.cell(240, 8, self._descricao, align='L')

        self.set_y(y_inicio + 28)
        self.set_text_color(0, 0, 0)

    def desenhar_cards_resumo(self, cards):
        """
        Desenha cartões informativos antes da tabela.
        cards: lista de dicts com {icone, valor, label}
        """
        if not cards:
            return

        y_inicio = self.get_y()
        num_cards = len(cards)
        largura_card = 55
        espaco = 5
        largura_total = num_cards * largura_card + (num_cards - 1) * espaco
        x_inicio = (297 - largura_total) / 2

        for i, card in enumerate(cards):
            x = x_inicio + i * (largura_card + espaco)

            # Fundo branco com borda dourada
            self.set_fill_color(*COR_BRANCO)
            self.set_draw_color(*COR_DOURADO)
            self.set_line_width(0.4)
            self.rect(x, y_inicio, largura_card, 24, 'FD')

            # Ícone dourado
            self.set_font('Helvetica', '', 12)
            self.set_text_color(*COR_DOURADO)
            self.set_xy(x + 3, y_inicio + 2)
            self.cell(10, 8, card.get('icone', ''), align='L')

            # Valor grande
            self.set_font('Helvetica', 'B', 16)
            self.set_text_color(*COR_AZUL_MARINHO)
            self.set_xy(x + 12, y_inicio + 1)
            self.cell(largura_card - 15, 10, str(card.get('valor', 0)), align='L')

            # Label
            self.set_font('Helvetica', '', 7)
            self.set_text_color(*COR_CINZA_TEXTO)
            self.set_xy(x + 3, y_inicio + 14)
            self.cell(largura_card - 6, 8, card.get('label', ''), align='L')

        self.set_y(y_inicio + 30)
        self.set_text_color(0, 0, 0)

    def cabecalho_tabela(self, colunas, larguras):
        """Cabeçalho da tabela estilo profissional."""
        y = self.get_y()

        # Fundo azul-marinho
        self.set_fill_color(*COR_AZUL_MARINHO)
        self.set_draw_color(*COR_CINZA_LINHA)
        self.set_line_width(0.2)

        # Altura do cabeçalho
        alt_header = 6
        self.rect(12, y, 273, alt_header, 'F')

        # Textos do cabeçalho
        self.set_font('Helvetica', 'B', 7)
        self.set_text_color(*COR_BRANCO)

        x = 12
        num_colunas = min(len(colunas), len(larguras))
        for i in range(num_colunas):
            self.set_xy(x, y)
            self.cell(larguras[i], alt_header, colunas[i], border=0, align='C')
            x += larguras[i]

        self.set_y(y + alt_header)
        self.set_text_color(0, 0, 0)

    def linha_tabela(self, valores, larguras, idx_linha=0, cor_especial=None):
        """Gera uma linha da tabela com alternância de cores."""
        y = self.get_y()
        alt_linha = 6

        # Cor de fundo alternada
        if cor_especial:
            self.set_fill_color(*cor_especial)
        elif idx_linha % 2 == 0:
            self.set_fill_color(*COR_BRANCO)
        else:
            self.set_fill_color(*COR_BEGE)

        self.set_draw_color(*COR_CINZA_LINHA)
        self.set_line_width(0.15)

        # Fundo da linha
        self.rect(12, y, 273, alt_linha, 'F')

        # Textos
        self.set_font('Helvetica', '', 7)
        self.set_text_color(40, 40, 40)

        x = 12
        num_colunas = min(len(valores), len(larguras))
        for i in range(num_colunas):
            texto = str(valores[i]) if valores[i] is not None else '-'
            # Trunca texto
            max_chars = int(larguras[i] / 2.2)
            if len(texto) > max_chars:
                texto = texto[:max_chars - 1] + '...'

            self.set_xy(x, y)
            self.cell(larguras[i], alt_linha, texto, border=0, align='C')
            x += larguras[i]

        self.set_y(y + alt_linha)
        self.set_text_color(0, 0, 0)

    def caixa_observacao(self, texto):
        """Desenha caixa de observação no final do relatório."""
        y = self.get_y() + 5

        # Fundo bege
        self.set_fill_color(*COR_BEGE)
        self.set_draw_color(*COR_DOURADO)
        self.set_line_width(0.4)

        # Calcula altura necessária
        self.set_font('Helvetica', '', 8)
        linhas = self._calcular_linhas(texto, 170)
        alt_caixa = max(16, linhas * 5 + 10)

        self.rect(12, y, 273, alt_caixa, 'FD')

        # Icone de informacao
        self.set_font('Helvetica', 'B', 10)
        self.set_text_color(*COR_DOURADO)
        self.set_xy(16, y + 3)
        self.cell(6, 5, '[i]', align='L')

        # Texto "Observacao:"
        self.set_font('Helvetica', 'B', 8)
        self.set_text_color(*COR_AZUL_MARINHO)
        self.set_xy(22, y + 3)
        self.cell(25, 5, 'Observacao:', align='L')

        # Texto da observação
        self.set_font('Helvetica', '', 8)
        self.set_text_color(*COR_CINZA_TEXTO)
        self.set_xy(22, y + 9)
        self.multi_cell(170, 4, texto)

        self.set_y(y + alt_caixa + 5)
        self.set_text_color(0, 0, 0)

    def _calcular_linhas(self, texto, largura_max):
        """Calcula número de linhas necessárias para o texto."""
        palavras = texto.split()
        linha_atual = ''
        linhas = 1
        for palavra in palavras:
            teste = f"{linha_atual} {palavra}".strip()
            if self.get_string_width(teste) < largura_max:
                linha_atual = teste
            else:
                linhas += 1
                linha_atual = palavra
        return linhas

    def footer(self):
        """Rodé profissional com paginação e linha dourada."""
        # Linha dourada
        self.set_draw_color(*COR_DOURADO)
        self.set_line_width(0.5)
        self.line(12, self.get_y() + 2, 285, self.get_y() + 2)

        # Texto do rodape
        self.set_y(-18)
        self.set_font('Helvetica', 'I', 7)
        self.set_text_color(*COR_CINZA_TEXTO)
        self.cell(0, 5, f'{SCHOOL_NAME} - Sistema {SISTEMA_NAME}',
                  new_x="LMARGIN", new_y="NEXT", align='L')
        self.cell(0, 5, f'Pagina {self.page_no()}/{{nb}}',
                  new_x="LMARGIN", new_y="NEXT", align='R')
        self.set_text_color(0, 0, 0)


# ═══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO GENÉRICA DE GERAÇÃO DE PDF
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_pdf_generico(titulo_relatorio, colunas, larguras, dados, caminho,
                       resumo=None, descricao='', icone='', observacao='',
                       cards_resumo=None):
    """
    Gera PDF profissional para qualquer relatório.

    Args:
        titulo_relatorio: Título do relatório
        colunas: Lista de nomes das colunas
        larguras: Lista de larguras das colunas
        dados: Lista de tuplas com os dados
        caminho: Caminho do arquivo de saída
        resumo: Dict com totais (opcional)
        descricao: Descrição do relatório (opcional)
        icone: Ícone Unicode do relatório (opcional)
        observacao: Texto da observação (opcional)
        cards_resumo: Lista de dicts para cards (opcional)
    """
    # Mapeia título para ícone se não fornecido
    if not icone:
        titulo_lower = titulo_relatorio.lower()
        for chave, valor in ICONES_RELATORIOS.items():
            if chave in titulo_lower:
                icone = valor
                break

    pdf = RelatorioPDF(titulo_relatorio, descricao, icone)
    pdf.alias_nb_pages()
    pdf.add_page()

    # Tabela
    if not dados:
        pdf.set_font('Helvetica', '', 11)
        pdf.set_text_color(*COR_CINZA_TEXTO)
        pdf.ln(10)
        pdf.cell(0, 8, 'Nenhum registro encontrado.', new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.set_text_color(0, 0, 0)
    else:
        pdf.cabecalho_tabela(colunas, larguras)
        for i, linha in enumerate(dados):
            cor = None
            # Alerta para atrasos > 30 dias
            if len(linha) > 6 and isinstance(linha[6], (int, float)) and linha[6] > 30:
                cor = COR_VERMELHO_SUAVE
            pdf.linha_tabela(linha, larguras, idx_linha=i, cor_especial=cor)

            # Verifica espaço na página
            if pdf.get_y() > 175:
                pdf.add_page()
                pdf.cabecalho_tabela(colunas, larguras)

        # Total de registros
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(*COR_AZUL_MARINHO)
        pdf.cell(0, 6, f'Total de registros: {len(dados)}',
                 new_x="LMARGIN", new_y="NEXT", align='R')
        pdf.set_text_color(0, 0, 0)

    # Caixa de observação
    if observacao:
        pdf.caixa_observacao(observacao)

    pdf.output(caminho)
    return caminho


# ═══════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_excel_generico(titulo_relatorio, colunas, dados, caminho, resumo=None):
    """Gera Excel genérico para qualquer relatório."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo_relatorio[:31]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
    header_fill = PatternFill(start_color="0B1D34", end_color="0B1D34", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    titulo_cell = ws.cell(row=1, column=1, value=f"{SCHOOL_NAME} - {titulo_relatorio}")
    titulo_cell.font = Font(bold=True, size=14, name='Calibri', color="0B1D34")
    titulo_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(colunas))
    data_cell = ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    data_cell.font = Font(size=10, italic=True, name='Calibri', color="666666")
    data_cell.alignment = Alignment(horizontal="center")

    # Resumo
    row_start = 4
    if resumo:
        for chave, valor in resumo.items():
            ws.cell(row=row_start, column=1, value=chave).font = Font(bold=True, name='Calibri')
            ws.cell(row=row_start, column=2, value=valor)
            row_start += 1
        row_start += 1

    # Cabeçalho da tabela
    for i, col in enumerate(colunas, 1):
        cell = ws.cell(row=row_start, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados
    thin_border = Border(
        left=Side(style='thin', color='DCDCDC'),
        right=Side(style='thin', color='DCDCDC'),
        top=Side(style='thin', color='DCDCDC'),
        bottom=Side(style='thin', color='DCDCDC')
    )

    for r, linha in enumerate(dados, row_start + 1):
        for c, val in enumerate(linha, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Alternância de cores
            if (r - row_start) % 2 == 0:
                cell.fill = PatternFill(start_color="FAF7F2", end_color="FAF7F2", fill_type="solid")

    # Ajusta largura das colunas
    for i, col in enumerate(colunas, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = max(len(col) + 2, 15)

    # Total
    total_row = row_start + len(dados) + 2
    ws.cell(row=total_row, column=1, value=f"Total de registros: {len(dados)}").font = Font(bold=True, name='Calibri')

    wb.save(caminho)
    return caminho
